
import io, os, json, zipfile, datetime, traceback
from typing import Dict, List
import streamlit as st, pandas as pd
from openpyxl import load_workbook
from parsers import diesel_golden, ngk_golden, bosch_golden

PARSERS = {
    "auto": None,
    "diesel_golden": diesel_golden,
    "ngk_golden": ngk_golden,
    "bosch_golden": bosch_golden,
}

EXTRACTORS = ["auto", "pdfplumber", "pypdf", "pymupdf"]

st.set_page_config(page_title="Invoice PDF -> Excel (PLUS v3.3.20 clean)", layout="wide")
st.title("Invoice PDF -> Excel (PLUS v3.3.20 clean)")

with st.sidebar:
    override = st.selectbox("Vendor override", list(PARSERS.keys()), index=0)
    extractor = st.selectbox("Extractor", EXTRACTORS, index=0, help="Pin a specific text extractor if needed.")
    show_debug = st.checkbox("Show debug (raw text + parsed items)", value=False)
    st.markdown("---")
    if st.button("Run self-tests (goldens/)"):
        st.write("Running golden tests...")
        results = []
        try:
            cfg_path = os.path.join("goldens", "goldens.json")
            if not os.path.exists(cfg_path):
                st.warning("goldens/goldens.json not found. Add golden PDFs and config first.")
            else:
                with open(cfg_path, "r", encoding="utf-8") as fh:
                    cfg = json.load(fh)
                for test in cfg.get("tests", []):
                    fname = test.get("file")
                    path = os.path.join("goldens", fname)
                    if not os.path.exists(path):
                        results.append((fname, "MISSING_FILE", 0, "File not found"))
                        continue
                    with open(path, "rb") as fhpdf:
                        data = fhpdf.read()
                    text = st.session_state.get("extract_text_fn", lambda b: "")(data)
                    parser_key = test.get("override", "")
                    parser = PARSERS.get(parser_key)
                    if parser is None:
                        results.append((fname, "BAD_PARSER", 0, f"Unknown parser {parser_key}"))
                        continue
                    header, items = parser.parse(text)
        # valeo_dual_export
        packing_rows = []
        try:
            if vendor_key == 'valeo_golden':
                from parsers import valeo_golden as _val
                packing_rows = _val.parse_packing(text)
        except Exception as _e:
            packing_rows = []
                    ok = True
                    msg = []
                    exp = int(test.get("expected_count", 0))
                    if exp and len(items) != exp:
                        ok = False; msg.append(f"count {len(items)} != expected {exp}")
                    must = test.get("must_include_items", [])
                    if must:
                        got_items = {str(it.get('Item')) for it in items}
                        miss = [m for m in must if m not in got_items]
                        if miss:
                            ok = False; msg.append(f"missing items {miss}")
                    results.append((fname, "PASS" if ok else "FAIL", len(items), "; ".join(msg)))
        except Exception as e:
            st.error(f"Self-test error: {e}")
            st.text(traceback.format_exc())
            results = []
        if results:
            df = pd.DataFrame(results, columns=["file","result","parsed_items","notes"])
            st.dataframe(df, use_container_width=True)

def extract_text(file_bytes: bytes) -> str:
    # save chosen extractor function in session for self-tests
    def run_auto(data: bytes) -> str:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                texts = [p.extract_text() or "" for p in pdf.pages]
            txt = "\n".join(texts)
            if txt.strip(): return txt
        except: pass
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            texts = [p.extract_text() or "" for p in reader.pages]
            txt = "\n".join(texts)
            if txt.strip(): return txt
        except: pass
        try:
            import fitz
            doc = fitz.open(stream=data, filetype="pdf")
            texts = [p.get_text("text") or "" for p in doc]
            return "\n".join(texts)
        except: return ""
    def run_pdfplumber(data: bytes) -> str:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                return "\n".join([p.extract_text() or "" for p in pdf.pages])
        except: return ""
    def run_pypdf(data: bytes) -> str:
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            return "\n".join([p.extract_text() or "" for p in reader.pages])
        except: return ""
    def run_pymupdf(data: bytes) -> str:
        try:
            import fitz
            doc = fitz.open(stream=data, filetype="pdf")
            return "\n".join([p.get_text("text") or "" for p in doc])
        except: return ""

    table = {
        "auto": run_auto,
        "pdfplumber": run_pdfplumber,
        "pypdf": run_pypdf,
        "pymupdf": run_pymupdf,
    }
    fn = table.get(extractor or "auto", run_auto)
    st.session_state["extract_text_fn"] = fn
    return fn(file_bytes)

def parse_dispatch(text: str, override_key: str):
    if override_key and override_key != "auto":
        parser = PARSERS[override_key]
        used = override_key
        return PARSERS[override_key].parse(text), used
    used = "auto"
    if "Diesel Technic" in text:
        return diesel_golden.parse(text), "diesel_golden"
    if "NGK" in text or "Niterra" in text:
        return ngk_golden.parse(text), "ngk_golden"
    return ({}, []), used

def to_xlsx_bytes(header: Dict, items: List[Dict]) -> bytes:
    hdr = dict(header or {})
    hdr["Parsed on"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        pd.DataFrame([hdr]).to_excel(xw, sheet_name="Header", index=False)
        cols = ["Item", "Qty", "UoM", "Unit Price", "Amount", "VAT"]
        df = pd.DataFrame(items or [], columns=cols)
        # cast numerics and report any NaNs
        for c in ["Qty","Unit Price","Amount"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        df.to_excel(xw, sheet_name="ERP_Import", index=False)
    out.seek(0)
    wb = load_workbook(out)
    wb.active = wb["ERP_Import"]
    out2 = io.BytesIO()
    wb.save(out2)
    return out2.getvalue()

files = st.file_uploader("Upload one or more PDF files", type=["pdf"], accept_multiple_files=True)

if files:
    zip_buf = io.BytesIO()
    zf = zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED)
    for f in files:
        st.subheader(f.name)
        data = f.read()
        text = extract_text(data)
        (header, items), used_parser = parse_dispatch(text, override)

        # status line
        st.markdown(f"**Parser:** `{used_parser}`  •  **Extractor:** `{extractor}`  •  **Parsed items:** **{len(items)}**")

        # preview
        c1, c2 = st.columns(2)
        with c1:
            st.write("Header")
            st.dataframe(pd.DataFrame([header]) if header else pd.DataFrame(), use_container_width=True)
        with c2:
            st.write("ERP_Import")
            cols = ["Item", "Qty", "UoM", "Unit Price", "Amount", "VAT"]
            df = pd.DataFrame(items or [], columns=cols)
            st.dataframe(df, use_container_width=True, height=300)

        # numeric validation
        if df.shape[0] and df[["Qty","Unit Price","Amount"]].isna().any().any():
            bad = df[df[["Qty","Unit Price","Amount"]].isna().any(axis=1)]
            st.warning(f"Found {bad.shape[0]} rows with non-numeric values in Qty/Unit Price/Amount. Check debug.")
        if show_debug:
            st.write("Debug (raw text first 200 lines)")
            st.text("\n".join((text or "").splitlines()[:200]))

        xbytes = to_xlsx_bytes(header, items)
        st.download_button(f"⬇️ Download XLSX — {f.name}.xlsx", data=xbytes, file_name=f"{f.name}.xlsx")
        zf.writestr(f"{f.name}.xlsx", xbytes)
        if vendor_key == "valeo_golden" and (packing_rows is not None):
            import io
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "PackingList"
            ws.append(["Parcel N°","Valeo Material N","Quantity"])
            for r in (packing_rows or []): ws.append([r.get("Parcel N°"), r.get("Valeo Material N"), r.get("Quantity")])
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            zf.writestr(f"{f.name}_packing.xlsx", buf.getvalue())
    zf.close()
    st.download_button("📦 Download all as ZIP", data=zip_buf.getvalue(), file_name="converted_invoices.zip")
else:
    st.info("Upload PDF invoices to convert them. For regression checks, put samples into ./goldens and click 'Run self-tests' in the sidebar.")
